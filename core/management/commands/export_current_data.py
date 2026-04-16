from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import models
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font

from core.models import (
    AuditLog,
    Branch,
    Customer,
    DailyReconciliation,
    Inventory,
    InventoryAdjustment,
    Order,
    OrderChangeRequest,
    OrderItem,
    Payment,
    PaymentReversal,
    Product,
    ProductCategory,
    Receipt,
    UserProfile,
)

User = get_user_model()


def _sheet_title(title: str) -> str:
    invalid = set(r'[]:*?/\\')
    clean = "".join(ch for ch in title if ch not in invalid).strip()
    return clean[:31] or "Sheet"


def _serialize_value(value):
    if value is None:
        return ""
    if isinstance(value, models.Model):
        return str(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list, tuple)):
        return str(value)
    if isinstance(value, models.fields.files.FieldFile):
        if not value:
            return ""
        try:
            return value.url
        except Exception:
            return value.name or ""
    return value


def _field_names_for_model(model: type[models.Model]) -> list[str]:
    names: list[str] = []
    for field in model._meta.fields:
        names.append(field.name)
    return names


def _row_for_instance(instance: models.Model, field_names: list[str]) -> list[object]:
    row: list[object] = []
    for field_name in field_names:
        row.append(_serialize_value(getattr(instance, field_name)))
    return row


class Command(BaseCommand):
    help = "Export the current database records to an Excel workbook with one worksheet per model."

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            default="exports/furnisync_current_data.xlsx",
            help="Output path for the generated Excel workbook.",
        )

    def handle(self, *args, **options):
        output_path = Path(options["output"])
        if not output_path.is_absolute():
            output_path = Path.cwd() / output_path
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        export_plan: list[tuple[str, models.QuerySet, list[str] | None]] = [
            (
                "Users",
                User.objects.all().order_by("id"),
                ["id", "username", "first_name", "last_name", "email", "is_active", "is_staff", "is_superuser", "last_login", "date_joined"],
            ),
            ("User Profiles", UserProfile.objects.select_related("user", "branch").all().order_by("id"), None),
            ("Branches", Branch.objects.all().order_by("id"), None),
            ("Customers", Customer.objects.select_related("branch").all().order_by("id"), None),
            ("Product Categories", ProductCategory.objects.all().order_by("id"), None),
            ("Products", Product.objects.select_related("category").all().order_by("id"), None),
            ("Inventory", Inventory.objects.select_related("product", "branch").all().order_by("id"), None),
            ("Orders", Order.objects.select_related("customer", "branch", "created_by", "last_modified_by", "assigned_collector", "approved_by").all().order_by("id"), None),
            ("Order Items", OrderItem.objects.select_related("order", "product").all().order_by("id"), None),
            ("Payments", Payment.objects.select_related("order", "branch", "collector", "manager_resolved_by").all().order_by("id"), None),
            ("Receipts", Receipt.objects.select_related("payment", "order", "branch", "collector").all().order_by("id"), None),
            ("Reconciliations", DailyReconciliation.objects.select_related("branch", "collector", "approved_by").all().order_by("id"), None),
            ("Audit Logs", AuditLog.objects.select_related("user").all().order_by("id"), None),
            ("Inventory Adjustments", InventoryAdjustment.objects.select_related("product", "branch", "created_by", "approved_by", "audit_log").all().order_by("id"), None),
            ("Order Change Requests", OrderChangeRequest.objects.select_related("order", "requested_by", "reviewed_by", "requested_assigned_collector").all().order_by("id"), None),
            ("Payment Reversals", PaymentReversal.objects.select_related("payment", "requested_by", "reviewed_by").all().order_by("id"), None),
        ]

        summary_rows: list[list[object]] = [
            ["Exported At", timezone.now()],
            ["Database", "default"],
            ["Workbook", output_path.name],
        ]

        for sheet_name, queryset, explicit_fields in export_plan:
            worksheet = workbook.create_sheet(title=_sheet_title(sheet_name))
            field_names = explicit_fields or _field_names_for_model(queryset.model)
            worksheet.append(field_names)
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            count = 0
            for instance in queryset.iterator():
                worksheet.append(_row_for_instance(instance, field_names))
                count += 1

            worksheet.freeze_panes = "A2"
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)

            summary_rows.append([sheet_name, count])

        summary_sheet = workbook.create_sheet(title="Summary", index=0)
        for row in summary_rows:
            summary_sheet.append(row)
        for cell in summary_sheet[1]:
            cell.font = Font(bold=True)
        summary_sheet.freeze_panes = "A2"
        summary_sheet.column_dimensions["A"].width = 28
        summary_sheet.column_dimensions["B"].width = 24

        workbook.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"Excel export created: {output_path}"))
